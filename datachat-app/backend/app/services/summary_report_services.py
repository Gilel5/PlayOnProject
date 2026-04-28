"""
Excel summary report generation service.

Produces a multi-sheet .xlsx workbook from the finance table using openpyxl.
The workbook has three sheets:
  1. Revenue Breakdown   — transaction type totals, new purchases, rebills, revenue share.
  2. Processor Transaction Pivot — base amount, processor fees, and pass counts by type.
  3. New Purchase by Day — daily purchase counts and revenue within the selected period.

Supports three report scopes:
  - annual      — all transactions in a given calendar year.
  - single_month— all transactions in one calendar month.
  - multimonth  — each month in a date range as a side-by-side column group.
"""

from __future__ import annotations

import io
from datetime import datetime

from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.db.finance_session import finance_engine
from app.services.openai_services import TABLE


# ── Date helpers ──────────────────────────────────────────────────────────────

def _month_label(month_str: str) -> str:
    """Convert a 'YYYY-MM' string to a human-readable label like 'Jan 2024'."""
    dt = datetime.strptime(month_str, "%Y-%m")
    return dt.strftime("%b %Y")


def _next_month(month_str: str) -> str:
    """Return the 'YYYY-MM' string for the month immediately following month_str."""
    dt = datetime.strptime(month_str, "%Y-%m")
    year = dt.year + (1 if dt.month == 12 else 0)
    month = 1 if dt.month == 12 else dt.month + 1
    return f"{year:04d}-{month:02d}"


def _month_range_list(start_month: str, end_month: str) -> list[str]:
    """
    Build an ordered list of 'YYYY-MM' strings from start_month to end_month
    inclusive. Used to iterate over months when building a multimonth report.
    """
    months = []
    current = start_month
    while current <= end_month:
        months.append(current)
        current = _next_month(current)
    return months


# ── Data fetching ─────────────────────────────────────────────────────────────

def _run_queries(conn, where_sql: str, params: dict) -> dict:
    """
    Execute all report queries for a single time period and return the results.

    All queries share the same WHERE clause (where_sql + params) so the same
    function works for annual, monthly, or any arbitrary date range.

    Returns a dict with keys:
      revenue_by_type   — totals by processor_transaction_type (Charge/Dispute/Fee/Refund)
      refund_stats      — aggregate refund amount and refund rate inputs
      new_purchase      — revenue and subscriber count by pass_name for new purchases
      rebill            — revenue and subscriber count by pass_name for rebills
      monthly_rebill    — rebill revenue broken down by billing_number
      revenue_share     — revenue and rev-share amount grouped by attribution_method
      processor_pivot   — base amount, processor fees, pass count by transaction type
      purchase_by_day   — daily new-purchase count and revenue
    """
    revenue_by_type = conn.execute(text(f"""
        SELECT
            processor_transaction_type,
            SUM(base_amount) AS revenue,
            COUNT(*) AS count
        FROM {TABLE}
        WHERE processor_transaction_type IN ('Charge','Dispute','Fee','Refund')
          AND {where_sql}
        GROUP BY processor_transaction_type
        ORDER BY
            CASE processor_transaction_type
                WHEN 'Charge' THEN 1
                WHEN 'Dispute' THEN 2
                WHEN 'Fee' THEN 3
                WHEN 'Refund' THEN 4
            END
    """), params).fetchall()

    refund_stats = conn.execute(text(f"""
        SELECT
            SUM(CASE WHEN processor_transaction_type = 'Refund' THEN base_amount ELSE 0 END),
            SUM(base_amount),
            COUNT(CASE WHEN processor_transaction_type = 'Refund' THEN 1 END),
            COUNT(*)
        FROM {TABLE}
        WHERE processor_transaction_type IN ('Charge','Dispute','Fee','Refund')
          AND {where_sql}
    """), params).fetchone()

    new_purchase = conn.execute(text(f"""
        SELECT COALESCE(pass_name, 'Null'), SUM(base_amount), COUNT(*)
        FROM {TABLE}
        WHERE transaction_type = 'New Purchase'
          AND {where_sql}
        GROUP BY pass_name
        ORDER BY CASE COALESCE(pass_name,'Null')
            WHEN 'Annual' THEN 1 WHEN 'Media' THEN 2
            WHEN 'Month' THEN 3 WHEN 'Null' THEN 4
            WHEN 'Season' THEN 5 ELSE 6 END
    """), params).fetchall()

    rebill = conn.execute(text(f"""
        SELECT COALESCE(pass_name, 'Null'), SUM(base_amount), COUNT(*)
        FROM {TABLE}
        WHERE transaction_type = 'Rebill'
          AND {where_sql}
        GROUP BY pass_name
        ORDER BY CASE COALESCE(pass_name,'Null')
            WHEN 'Annual' THEN 1 WHEN 'Media' THEN 2
            WHEN 'Month' THEN 3 WHEN 'Null' THEN 4
            WHEN 'Season' THEN 5 ELSE 6 END
    """), params).fetchall()

    monthly_rebill = conn.execute(text(f"""
        SELECT billing_number, SUM(base_amount), COUNT(*)
        FROM {TABLE}
        WHERE transaction_type = 'Rebill'
          AND billing_number IS NOT NULL
          AND {where_sql}
        GROUP BY billing_number
        ORDER BY billing_number
    """), params).fetchall()

    revenue_share = conn.execute(text(f"""
        SELECT COALESCE(attribution_method, 'Null'), SUM(base_amount), SUM(revenue_share_usd)
        FROM {TABLE}
        WHERE processor_transaction_type = 'Charge'
          AND {where_sql}
        GROUP BY attribution_method
        ORDER BY attribution_method
    """), params).fetchall()

    processor_pivot = conn.execute(text(f"""
        SELECT
            processor_transaction_type,
            SUM(base_amount),
            SUM(CASE WHEN processor_fee ~ '^[0-9.\\-]+$'
                     THEN CAST(processor_fee AS NUMERIC) ELSE 0 END),
            COUNT(pass_name)
        FROM {TABLE}
        WHERE processor_transaction_type IN ('Charge','Dispute','Fee','Refund')
          AND {where_sql}
        GROUP BY processor_transaction_type
        ORDER BY CASE processor_transaction_type
            WHEN 'Charge' THEN 1 WHEN 'Dispute' THEN 2
            WHEN 'Fee' THEN 3 WHEN 'Refund' THEN 4 END
    """), params).fetchall()

    purchase_by_day = conn.execute(text(f"""
        SELECT
            TO_CHAR(transaction_date, 'Mon DD'),
            COUNT(*),
            SUM(base_amount)
        FROM {TABLE}
        WHERE transaction_type = 'New Purchase'
          AND transaction_date IS NOT NULL
          AND {where_sql}
        GROUP BY DATE_TRUNC('day', transaction_date), TO_CHAR(transaction_date, 'Mon DD')
        ORDER BY DATE_TRUNC('day', transaction_date)
    """), params).fetchall()

    return {
        "revenue_by_type": revenue_by_type,
        "refund_stats": refund_stats,
        "new_purchase": new_purchase,
        "rebill": rebill,
        "monthly_rebill": monthly_rebill,
        "revenue_share": revenue_share,
        "processor_pivot": processor_pivot,
        "purchase_by_day": purchase_by_day,
    }


# ── Workbook builder ──────────────────────────────────────────────────────────

def generate_summary_reports(
    report_type: str,
    year: int | None = None,
    month: str | None = None,
    start_month: str | None = None,
    end_month: str | None = None,
) -> bytes:
    """
    Build a styled .xlsx workbook for the requested time period and return it
    as raw bytes suitable for streaming to the client.

    Args:
        report_type:  One of "annual", "single_month", or "multimonth".
        year:         Calendar year (annual only).
        month:        "YYYY-MM" string (single_month only).
        start_month:  "YYYY-MM" start of range (multimonth only).
        end_month:    "YYYY-MM" end of range, inclusive (multimonth only).

    Raises ValueError for invalid report_type or missing required parameters.
    """
    # ── Shared cell styles ────────────────────────────────────────────────────
    header_font = Font(bold=True, size=11)
    # header_fill and total_fill share the same colour — one definition used for both
    accent_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    total_font = Font(bold=True, size=11)
    section_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    month_font = Font(bold=True, size=12)
    currency_fmt = '#,##0.00'
    integer_fmt = '#,##0'
    pct_fmt = '0.0%'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, row, start_col, cols):
        """Apply bold + accent fill + border + center alignment to a header row."""
        for c in range(start_col, start_col + cols):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = accent_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    def style_row(ws, row, start_col, cols, is_total=False):
        """Apply border to a data row; optionally apply bold + accent fill for totals."""
        for c in range(start_col, start_col + cols):
            cell = ws.cell(row=row, column=c)
            cell.border = thin_border
            if is_total:
                cell.font = total_font
                cell.fill = accent_fill

    # ── Fetch data for each time period ──────────────────────────────────────
    with finance_engine.connect() as conn:
        if report_type == "annual":
            if not year:
                raise ValueError("year is required for annual reports")
            datasets = [{
                "label": str(year),
                "data": _run_queries(conn, "EXTRACT(YEAR FROM transaction_date) = :year", {"year": year}),
            }]
        elif report_type == "single_month":
            if not month:
                raise ValueError("month is required for single-month reports")
            next_month = _next_month(month)
            datasets = [{
                "label": _month_label(month),
                "data": _run_queries(
                    conn,
                    "transaction_date >= CAST(:month_start AS DATE) AND transaction_date < CAST(:next_month_start AS DATE)",
                    {"month_start": f"{month}-01", "next_month_start": f"{next_month}-01"},
                ),
            }]
        elif report_type == "multimonth":
            if not start_month or not end_month:
                raise ValueError("start_month and end_month are required for multi-month reports")
            month_list = _month_range_list(start_month, end_month)
            datasets = []
            for m in month_list:
                next_month = _next_month(m)
                datasets.append({
                    "label": _month_label(m),
                    "data": _run_queries(
                        conn,
                        "transaction_date >= CAST(:month_start AS DATE) AND transaction_date < CAST(:next_month_start AS DATE)",
                        {"month_start": f"{m}-01", "next_month_start": f"{next_month}-01"},
                    ),
                })
        else:
            raise ValueError("Invalid report_type")

    # ── Build workbook sheets ─────────────────────────────────────────────────
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Revenue Breakdown"
    ws2 = wb.create_sheet("Processor Transaction Pivot")
    ws3 = wb.create_sheet("New Purchase by Day")

    def write_revenue_block(ws, start_col, label, data):
        """Write the Revenue Breakdown section into ws starting at start_col."""
        for col in range(start_col, start_col + 3):
            ws.column_dimensions[chr(64 + col)].width = 22

        r = 1
        ws.cell(r, start_col, label).font = month_font
        r += 2

        revenue_by_type = data["revenue_by_type"]
        refund_stats = data["refund_stats"]
        new_purchase = data["new_purchase"]
        rebill = data["rebill"]
        monthly_rebill = data["monthly_rebill"]
        revenue_share = data["revenue_share"]

        ws.cell(r, start_col, "Type")
        ws.cell(r, start_col + 1, "Revenue")
        ws.cell(r, start_col + 2, "Count")
        style_header(ws, r, start_col, 3)
        r += 1

        for row in revenue_by_type:
            ws.cell(r, start_col, row[0])
            ws.cell(r, start_col + 1, float(row[1] or 0)).number_format = currency_fmt
            ws.cell(r, start_col + 2, int(row[2] or 0)).number_format = integer_fmt
            style_row(ws, r, start_col, 3)
            r += 1

        if refund_stats and refund_stats[1]:
            ws.cell(r, start_col, "Refund Rate")
            ws.cell(r, start_col + 1, float(refund_stats[0] or 0) / float(refund_stats[1])).number_format = pct_fmt
            style_row(ws, r, start_col, 3)
            r += 1

        dispute_rev = sum(row[1] or 0 for row in revenue_by_type if row[0] == "Dispute")
        refund_rev = sum(row[1] or 0 for row in revenue_by_type if row[0] == "Refund")
        dispute_cnt = sum(row[2] or 0 for row in revenue_by_type if row[0] == "Dispute")
        refund_cnt = sum(row[2] or 0 for row in revenue_by_type if row[0] == "Refund")

        ws.cell(r, start_col, "Total Refunds & Disputes")
        ws.cell(r, start_col + 1, float(dispute_rev + refund_rev)).number_format = currency_fmt
        ws.cell(r, start_col + 2, int(dispute_cnt + refund_cnt)).number_format = integer_fmt
        style_row(ws, r, start_col, 3)
        r += 1

        gross = sum(row[1] or 0 for row in revenue_by_type)
        ws.cell(r, start_col, "Gross Base Revenue")
        ws.cell(r, start_col + 1, float(gross)).number_format = currency_fmt
        style_row(ws, r, start_col, 3, is_total=True)
        r += 2

        ws.cell(r, start_col, "New Purchase").font = section_font
        r += 1
        ws.cell(r, start_col, "Pass Name")
        ws.cell(r, start_col + 1, "Revenue")
        ws.cell(r, start_col + 2, "New Subscribers")
        style_header(ws, r, start_col, 3)
        r += 1

        np_total_rev, np_total_sub = 0, 0
        for row in new_purchase:
            ws.cell(r, start_col, row[0])
            ws.cell(r, start_col + 1, float(row[1] or 0)).number_format = currency_fmt
            ws.cell(r, start_col + 2, int(row[2] or 0)).number_format = integer_fmt
            np_total_rev += float(row[1] or 0)
            np_total_sub += int(row[2] or 0)
            style_row(ws, r, start_col, 3)
            r += 1

        ws.cell(r, start_col, "Gross New Purchases")
        ws.cell(r, start_col + 1, np_total_rev).number_format = currency_fmt
        ws.cell(r, start_col + 2, np_total_sub).number_format = integer_fmt
        style_row(ws, r, start_col, 3, is_total=True)
        r += 2

        ws.cell(r, start_col, "Rebill").font = section_font
        r += 1
        ws.cell(r, start_col, "Pass Name")
        ws.cell(r, start_col + 1, "Revenue")
        ws.cell(r, start_col + 2, "Subscribers")
        style_header(ws, r, start_col, 3)
        r += 1

        rb_total_rev, rb_total_sub = 0, 0
        for row in rebill:
            ws.cell(r, start_col, row[0])
            ws.cell(r, start_col + 1, float(row[1] or 0)).number_format = currency_fmt
            ws.cell(r, start_col + 2, int(row[2] or 0)).number_format = integer_fmt
            rb_total_rev += float(row[1] or 0)
            rb_total_sub += int(row[2] or 0)
            style_row(ws, r, start_col, 3)
            r += 1

        ws.cell(r, start_col, "Gross Rebills")
        ws.cell(r, start_col + 1, rb_total_rev).number_format = currency_fmt
        ws.cell(r, start_col + 2, rb_total_sub).number_format = integer_fmt
        style_row(ws, r, start_col, 3, is_total=True)
        r += 2

        ws.cell(r, start_col, "Monthly Rebill Revenue").font = section_font
        r += 1
        ws.cell(r, start_col, "Billing #")
        ws.cell(r, start_col + 1, "Revenue")
        ws.cell(r, start_col + 2, "Subscribers")
        style_header(ws, r, start_col, 3)
        r += 1

        for row in monthly_rebill:
            ws.cell(r, start_col, float(row[0]) if row[0] is not None else "")
            ws.cell(r, start_col + 1, float(row[1] or 0)).number_format = currency_fmt
            ws.cell(r, start_col + 2, int(row[2] or 0)).number_format = integer_fmt
            style_row(ws, r, start_col, 3)
            r += 1

        r += 1

        if revenue_share:
            ws.cell(r, start_col, "Revenue Share").font = section_font
            r += 1
            ws.cell(r, start_col, "Attribution")
            ws.cell(r, start_col + 1, "Revenue")
            ws.cell(r, start_col + 2, "Rev Share")
            style_header(ws, r, start_col, 3)
            r += 1

            for row in revenue_share:
                ws.cell(r, start_col, row[0])
                ws.cell(r, start_col + 1, float(row[1] or 0)).number_format = currency_fmt
                ws.cell(r, start_col + 2, float(row[2] or 0)).number_format = currency_fmt
                style_row(ws, r, start_col, 3)
                r += 1

    def write_processor_block(ws, start_col, label, data):
        """Write the Processor Transaction Pivot section into ws starting at start_col."""
        for col in range(start_col, start_col + 4):
            ws.column_dimensions[chr(64 + col)].width = 22

        r = 1
        ws.cell(r, start_col, label).font = month_font
        r += 2

        ws.cell(r, start_col, "Processor Transaction Type")
        ws.cell(r, start_col + 1, "Sum of Base Amount")
        ws.cell(r, start_col + 2, "Sum of Processor Fee")
        ws.cell(r, start_col + 3, "Count of Pass Name")
        style_header(ws, r, start_col, 4)
        r += 1

        gt_base, gt_fee, gt_count = 0, 0, 0
        for row in data["processor_pivot"]:
            ws.cell(r, start_col, row[0])
            ws.cell(r, start_col + 1, float(row[1] or 0)).number_format = currency_fmt
            ws.cell(r, start_col + 2, float(row[2] or 0)).number_format = currency_fmt
            ws.cell(r, start_col + 3, int(row[3] or 0)).number_format = integer_fmt
            gt_base += float(row[1] or 0)
            gt_fee += float(row[2] or 0)
            gt_count += int(row[3] or 0)
            style_row(ws, r, start_col, 4)
            r += 1

        ws.cell(r, start_col, "Grand Total")
        ws.cell(r, start_col + 1, gt_base).number_format = currency_fmt
        ws.cell(r, start_col + 2, gt_fee).number_format = currency_fmt
        ws.cell(r, start_col + 3, gt_count).number_format = integer_fmt
        style_row(ws, r, start_col, 4, is_total=True)

    def write_purchase_day_block(ws, start_col, label, data):
        """Write the New Purchase by Day section into ws starting at start_col."""
        for col in range(start_col, start_col + 3):
            ws.column_dimensions[chr(64 + col)].width = 18

        r = 1
        ws.cell(r, start_col, label).font = month_font
        r += 2

        ws.cell(r, start_col, "Day")
        ws.cell(r, start_col + 1, "Count of Transaction Type")
        ws.cell(r, start_col + 2, "Sum of Base Amount")
        style_header(ws, r, start_col, 3)
        r += 1

        gt_cnt, gt_amt = 0, 0
        for row in data["purchase_by_day"]:
            ws.cell(r, start_col, row[0])
            ws.cell(r, start_col + 1, int(row[1] or 0)).number_format = integer_fmt
            ws.cell(r, start_col + 2, float(row[2] or 0)).number_format = currency_fmt
            gt_cnt += int(row[1] or 0)
            gt_amt += float(row[2] or 0)
            style_row(ws, r, start_col, 3)
            r += 1

        ws.cell(r, start_col, "Grand Total")
        ws.cell(r, start_col + 1, gt_cnt).number_format = integer_fmt
        ws.cell(r, start_col + 2, gt_amt).number_format = currency_fmt
        style_row(ws, r, start_col, 3, is_total=True)

    ws1.cell(1, 1, "Revenue Breakdown").font = title_font
    ws2.cell(1, 1, "Processor Transaction Pivot").font = title_font
    ws3.cell(1, 1, "New Purchase by Day").font = title_font

    if report_type == "multimonth":
        # Each month occupies a fixed-width column group; groups sit side by side
        revenue_width = 4
        processor_width = 5
        purchase_width = 4

        for i, dataset in enumerate(datasets):
            write_revenue_block(ws1, 1 + i * revenue_width, dataset["label"], dataset["data"])
            write_processor_block(ws2, 1 + i * processor_width, dataset["label"], dataset["data"])
            write_purchase_day_block(ws3, 1 + i * purchase_width, dataset["label"], dataset["data"])
    else:
        dataset = datasets[0]
        write_revenue_block(ws1, 1, dataset["label"], dataset["data"])
        write_processor_block(ws2, 1, dataset["label"], dataset["data"])
        write_purchase_day_block(ws3, 1, dataset["label"], dataset["data"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()